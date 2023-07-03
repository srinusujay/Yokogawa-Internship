from openpyxl import Workbook, load_workbook

wb = load_workbook('edit.xlsx')
ws = wb.active
s = ws['B7'].value

cha = "A"
chb = "B"
chc = "C"
enter = "D"

lookup = {
	"TRANSMITTER" : "TX",
	"PRESSURE" : "PR",
	"FLOW" : "FL" ,
	"TEMPERATURE" : "TM",
	"ELEMENT" : "EL",
	"REACTOR" : "RCT"
}

for i in range(2, 5260):
	# INITIAL SETUP-------------
	cha = cha + str(i)
	chb = chb + str(i)
	chc = chc + str(i)
	enter = enter + str(i)
	loop = str(ws[cha].value) 
	if loop == "None":
		loop = ""
	comm = str(ws[chb].value)
	if comm == "None":
		comm = ""
	inst = str(ws[chc].value)
	if inst == "None":
		inst = ""

	# LOOP SETUP-----------------

	loop = loop.replace("-", "")

	# INSTRU SETUP---------------

	resi = inst.split()
	tempi = []
	for j in resi:
		tempi.append(j[0:2])
	print(tempi)


	# COMM SETUP-----------------

	comm = comm.replace("-", "")
	comm = comm.replace("/", "")
	comm = comm.replace("+", "")
	comm = comm.replace(".", "")
	comm = comm.replace("", "")

	comm = comm.replace("FROM", "")
	comm = comm.replace("TO", "")

	res = comm.split()
	newres = []
	count = 0

	for e in res:
		count = 0

		if count < 13:
			if e in resi:
				pass
			elif e in lookup.keys():
				newres.append(lookup[e])
			elif len(e) > 2:
				f, l = e[0], e[-1]
				e = f + l
				newres.append(e)
				count += len(e)
			elif len(e) == 1 :
				newres.append(e)
				count += len(e)

		print(e)
	print(newres)
	

	# FINAL CONVR----------------
	comm = " ".join(newres)
	inst = " ".join(tempi) 
	loop = " " + loop + " "
	final = inst + loop + comm
	final = final.upper()
	ws[enter].value = final

	# DEFAULT VALUE--------------
	print(final)
	cha = "A"
	chb = "B"
	chc = "C"
	enter = "D"

wb.save('edit.xlsx')
print(s[::-1])




































